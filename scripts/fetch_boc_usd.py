import csv
import json
import os
import re
import time
from datetime import datetime
from typing import Dict, List, Tuple, Optional

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

BOC_URL = "https://www.bankofchina.com/sourcedb/whpj/enindex_1619.html"
TIMEOUT = 30
DATA_DIR = "data"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

def ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)

def normalize_spaces(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())

def try_parse_pub_time(pub_time_raw: str) -> Tuple[str, str]:
    """
    Accepts multiple formats:
    - YYYY/MM/DD HH:MM:SS
    - YYYY-MM-DD HH:MM:SS
    Returns (date_iso, dt_iso).
    """
    s = normalize_spaces(pub_time_raw)
    s = s.replace("\u00a0", " ")  # non-breaking space
    # Common variants
    fmts = [
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d %H:%M",
    ]
    for f in fmts:
        try:
            dt = datetime.strptime(s, f)
            return dt.strftime("%Y-%m-%d"), dt.strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            pass
    raise ValueError(f"Unrecognized Pub Time format: '{pub_time_raw}'")

def fetch_html_with_retries(url: str, retries: int = 3, sleep_s: int = 2) -> str:
    last_err = None
    for attempt in range(1, retries + 1):
        try:
            resp = requests.get(url, timeout=TIMEOUT, headers=HEADERS)
            resp.raise_for_status()
            return resp.text
        except Exception as e:
            last_err = e
            print(f"[WARN] Fetch attempt {attempt}/{retries} failed: {e}")
            if attempt < retries:
                time.sleep(sleep_s)
    raise RuntimeError(f"Failed to fetch after {retries} attempts: {last_err}")

def find_usd_row(soup: BeautifulSoup) -> List[str]:
    """
    Searches all tables and returns the USD row as list of cell texts.
    Expect row like:
    USD | Buying | Cash Buying | Selling | Cash Selling | Middle | Pub Time
    """
    tables = soup.find_all("table")
    print(f"[INFO] Found {len(tables)} tables on page.")
    if not tables:
        raise RuntimeError("No tables found on BOC page.")

    # Try to find a row where first cell == USD (case-insensitive)
    for ti, table in enumerate(tables, start=1):
        rows = table.find_all("tr")
        for ri, tr in enumerate(rows, start=1):
            tds = tr.find_all("td")
            if not tds:
                continue
            first = normalize_spaces(tds[0].get_text())
            if first.upper() == "USD":
                cols = [normalize_spaces(td.get_text()) for td in tds]
                print(f"[INFO] USD row found in table {ti}, row {ri}: {cols}")
                return cols

    # Fallback: sometimes currency might be shown as "U.S. Dollar" in some versions
    for ti, table in enumerate(tables, start=1):
        rows = table.find_all("tr")
        for ri, tr in enumerate(rows, start=1):
            tds = tr.find_all("td")
            if not tds:
                continue
            first = normalize_spaces(tds[0].get_text()).lower()
            if "usd" == first or "u.s." in first or "dollar" in first:
                cols = [normalize_spaces(td.get_text()) for td in tds]
                # ensure USD appears somewhere
                if cols and (cols[0].upper() == "USD" or "USD" in cols[0].upper()):
                    print(f"[INFO] USD-like row found in table {ti}, row {ri}: {cols}")
                    return cols

    raise RuntimeError("USD row not found in any table.")

def fetch_usd_record() -> Dict[str, str]:
    html = fetch_html_with_retries(BOC_URL, retries=3, sleep_s=2)
    soup = BeautifulSoup(html, "html.parser")

    cols = find_usd_row(soup)

    # Guard: some pages add extra columns; we pick by position based on expected layout
    # 0 Currency
    # 1 Buying
    # 2 CashBuying
    # 3 Selling
    # 4 CashSelling
    # 5 Middle
    # 6 PubTime
    def safe(i: int) -> str:
        return cols[i] if i < len(cols) else ""

    currency = safe(0).upper()
    buying = safe(1)
    cash_buying = safe(2)
    selling = safe(3)
    cash_selling = safe(4)
    middle = safe(5)
    pub_time_raw = safe(6)

    if currency != "USD":
        raise RuntimeError(f"Row currency mismatch. Got: {currency}")

    if not pub_time_raw:
        # some layouts might place Pub Time elsewhere; try last column
        pub_time_raw = cols[-1] if cols else ""
    if not pub_time_raw:
        raise RuntimeError("Pub Time missing in USD row.")

    date_iso, pub_time_iso = try_parse_pub_time(pub_time_raw)

    record = {
        "date": date_iso,
        "publishTime": pub_time_iso,
        "publishTimeRaw": pub_time_raw,
        "currency": "USD",
        "buying": buying,
        "cashBuying": cash_buying,
        "selling": selling,
        "cashSelling": cash_selling,
        "middle": middle,
        "source": BOC_URL,
        "capturedAtUtc": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
    }
    print(f"[INFO] Parsed record: {record}")
    return record

def month_paths(date_iso: str) -> Tuple[str, str, str, str]:
    yyyy = date_iso[:4]
    mm = date_iso[5:7]
    folder = os.path.join(DATA_DIR, yyyy)
    ensure_dir(folder)
    base = os.path.join(folder, f"{yyyy}-{mm}")
    return folder, base + ".json", base + ".csv", base + ".xlsx"

def load_json(path: str) -> List[Dict[str, str]]:
    if not os.path.exists(path):
        return []
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def save_json(path: str, data: List[Dict[str, str]]) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def save_csv(path: str, data: List[Dict[str, str]]) -> None:
    fieldnames = [
        "date", "publishTime", "publishTimeRaw", "currency",
        "buying", "cashBuying", "selling", "cashSelling", "middle",
        "source", "capturedAtUtc"
    ]
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in data:
            w.writerow({k: r.get(k, "") for k in fieldnames})

def autosize(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

def to_float(s: str) -> Optional[float]:
    try:
        return float(str(s).strip())
    except:
        return None

def build_daily_tables(all_rows: List[Dict[str, str]]):
    by_date = {}
    for r in all_rows:
        by_date.setdefault(r["date"], []).append(r)
    dates = sorted(by_date.keys())

    avg_header = ["date","publishes","avgBuying","avgCashBuying","avgSelling","avgCashSelling","avgMiddle","minMiddle","maxMiddle"]
    avg_rows = [avg_header]

    first_header = ["date","firstPublishTime","buying","cashBuying","selling","cashSelling","middle","publishTimeRaw"]
    first_rows = [first_header]

    for d in dates:
        rows = sorted(by_date[d], key=lambda x: x.get("publishTime",""))
        publishes = len(rows)

        fr = rows[0]
        first_rows.append([
            d, fr.get("publishTime",""),
            fr.get("buying",""), fr.get("cashBuying",""),
            fr.get("selling",""), fr.get("cashSelling",""),
            fr.get("middle",""), fr.get("publishTimeRaw","")
        ])

        def avg_of(key: str) -> str:
            vals = [to_float(r.get(key,"")) for r in rows]
            vals = [v for v in vals if v is not None]
            if not vals:
                return ""
            return f"{sum(vals)/len(vals):.4f}"

        mids = [to_float(r.get("middle","")) for r in rows]
        mids = [m for m in mids if m is not None]

        avg_rows.append([
            d, str(publishes),
            avg_of("buying"), avg_of("cashBuying"),
            avg_of("selling"), avg_of("cashSelling"),
            avg_of("middle"),
            (f"{min(mids):.4f}" if mids else ""),
            (f"{max(mids):.4f}" if mids else "")
        ])

    return avg_rows, first_rows

def save_xlsx(path: str, all_rows: List[Dict[str, str]]) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "All Published Values"

    header = ["date","publishTime","buying","cashBuying","selling","cashSelling","middle","publishTimeRaw","capturedAtUtc","source"]
    ws1.append(header)
    for r in all_rows:
        ws1.append([
            r.get("date",""), r.get("publishTime",""),
            r.get("buying",""), r.get("cashBuying",""),
            r.get("selling",""), r.get("cashSelling",""),
            r.get("middle",""), r.get("publishTimeRaw",""),
            r.get("capturedAtUtc",""), r.get("source","")
        ])
    autosize(ws1)

    ws2 = wb.create_sheet("Daily Summary")
    avg_table, first_table = build_daily_tables(all_rows)

    ws2.append(["Day Averages"])
    for row in avg_table:
        ws2.append(row)

    ws2.append([])
    ws2.append(["Day First Published Values"])
    for row in first_table:
        ws2.append(row)

    autosize(ws2)
    wb.save(path)

def dedupe_and_append(existing: List[Dict[str, str]], new_record: Dict[str, str]):
    key = f"{new_record.get('currency','')}|{new_record.get('publishTime','')}"
    seen = set(f"{r.get('currency','')}|{r.get('publishTime','')}" for r in existing)
    if key in seen:
        return existing, False
    existing.append(new_record)
    existing.sort(key=lambda x: x.get("publishTime",""))
    return existing, True

def main():
    rec = fetch_usd_record()
    _, json_path, csv_path, xlsx_path = month_paths(rec["date"])

    existing = load_json(json_path)
    updated, changed = dedupe_and_append(existing, rec)

    if not changed:
        print("[INFO] No new publishTime. Nothing to update.")
        return

    save_json(json_path, updated)
    save_csv(csv_path, updated)
    save_xlsx(xlsx_path, updated)

    print("[OK] Updated files:")
    print(" -", json_path)
    print(" -", csv_path)
    print(" -", xlsx_path)

if __name__ == "__main__":
    main()
